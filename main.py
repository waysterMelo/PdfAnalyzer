import hashlib
import platform
import hmac
import io
import json
import os
import queue
import re
import shutil
import subprocess
import tempfile
import threading
import tkinter as tk
from tkinter import ttk, filedialog
from openpyxl.worksheet.table import Table, TableStyleInfo
import cv2
import fitz
import numpy as np
import pandas as pd
from PIL import Image, ImageTk, ImageFilter, ImageEnhance
from datetime import datetime, timedelta
import sys
import pytesseract
from tkinter import messagebox
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from pytesseract import pytesseract
from ttkthemes.themed_tk import ThemedTk

# Chave secreta usada para a assinatura HMAC
SECRET_KEY = b"waystermelo@"
# Nome do arquivo de licença
LICENSE_FILE = "license.txt"

# Caminho para a imagem do logo
image_path = os.path.join(os.path.dirname(__file__), 'img', 'logo.webp')

def create_signature(data):
    """Cria uma assinatura HMAC-SHA256 para os dados fornecidos."""
    return hmac.new(SECRET_KEY, data.encode('utf-8'), hashlib.sha256).hexdigest()

def verify_signature(data, signature):
    """Verifica se a assinatura fornecida corresponde ao conteúdo dos dados."""
    return hmac.compare_digest(create_signature(data), signature)

def check_license():
    """Verifica se a licença de teste ainda é válida e inicializa a data de instalação se necessário."""
    global activation_time

    try:
        # Verificar se o arquivo de licença existe
        if os.path.exists(LICENSE_FILE):
            with open(LICENSE_FILE, "r") as file:
                content = json.load(file)
                # Cria uma string contendo a data de ativação e a duração
                data_str = f"{content['activation_time']}|{content['duracao']}"
                saved_signature = content.get("signature")
                duracao_em_minutos = content.get("duracao")  # Carregar a duração diretamente do arquivo de licença

                # Verificar se a duração foi corretamente especificada
                if not isinstance(duracao_em_minutos, int) or duracao_em_minutos <= 0:
                    raise ValueError("Duração da licença inválida ou ausente no arquivo de licença.")

                # Verificar a integridade do arquivo com a assinatura
                if not verify_signature(data_str, saved_signature):
                    messagebox.showerror("Erro de Licença", "A licença foi manipulada! O programa será encerrado.")
                    return False

                # Carregar a data de ativação a partir do arquivo
                activation_time = datetime.fromisoformat(content["activation_time"])
        else:
            messagebox.showerror("Erro de Licença", "Arquivo de licença não encontrado! O programa será encerrado.")
            return False

        # Verificar o tempo de expiração da licença
        expiration_time = activation_time + timedelta(minutes=duracao_em_minutos)
        if datetime.now() > expiration_time:
            # Se o período de teste tiver expirado
            messagebox.showwarning("Licença Expirada", "Seu período de teste expirou. O programa será encerrado.")
            return False
        else:
            # Calcula o tempo restante da licença
            remaining_time = expiration_time - datetime.now()
            remaining_minutes = int(remaining_time.total_seconds() // 60)
            messagebox.showinfo("Licença de Teste", f"Você tem {remaining_minutes} minutos restantes de teste.")
            return True

    except Exception as e:
        # Caso ocorra qualquer erro ao verificar a licença
        messagebox.showerror("Erro de Licença", f"Ocorreu um erro ao verificar a licença: {str(e)}")
        return False

def configurar_tesseract():
    """Configuração do Tesseract OCR."""
    # Define o caminho do diretório tessdata e o executável do Tesseract
    tessdata_prefix = r'C:/Program Files/Tesseract-OCR/tessdata/'
    tesseract_cmd = r'C:/Program Files/Tesseract-OCR/tesseract.exe'
    # Cria uma configuração para o Tesseract
    tesseract_config = TesseractConfig(tessdata_prefix, tesseract_cmd)
    # Testa a configuração para garantir que tudo está correto
    tesseract_config.test_setup()

def iniciar_interface_principal():
    """Inicia a interface principal da aplicação."""
    global root, bg_image
    # Verificar se a licença está válida antes de iniciar
    if not check_license():
        # Se a licença estiver expirada ou com erro, o programa é encerrado
        return

    # Inicializando a janela principal
    root = tk.Tk()
    root.title("PDF Analyzer Blank")
    root.configure(bg="#1C2833")
    root.resizable(True, True)  # Permitir redimensionamento da janela

    # Centralizando a janela na tela
    window_width, window_height = 900, 600
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    position_top = int((screen_height - window_height) / 2)
    position_left = int((screen_width - window_width) / 2)
    root.geometry(f"{window_width}x{window_height}+{position_left}+{position_top}")

    # Carregando a imagem de fundo
    try:
        image = Image.open(image_path)
        # Redimensiona a imagem para o tamanho da janela
        image = image.resize((window_width, window_height))
        bg_image = ImageTk.PhotoImage(image)
    except Exception as e:
        # Caso haja erro ao carregar a imagem
        messagebox.showerror("Erro", f"Erro ao carregar a imagem de fundo: {e}")
        return

    # Rótulo para a imagem de fundo
    background_label = tk.Label(root, image=bg_image)
    background_label.place(relwidth=1, relheight=1)

    # Nome da aplicação no centro da tela
    app_name = tk.Label(root, text="PDF Analyzer Blank", font=("Helvetica", 40, "bold"),
                        fg="#FFFFFF", bg="#1E3D59", padx=20, pady=10, relief="raised", bd=10)
    app_name.place(relx=0.5, rely=0.25, anchor='center')

    # Função para iniciar a análise e carregar GUI principal
    def iniciar_analise():
        # Destroi a janela inicial e inicia a GUI do analisador
        root.destroy()
        configurar_tesseract()
        PDFAnalyzerGUI()

    # Botão para iniciar a análise do PDF
    start_button = tk.Button(root, text="Iniciar Análise", font=("Helvetica", 18, "bold"),
                             fg="#FFFFFF", bg="#1E3D59", activebackground="#34495E",
                             activeforeground="#FFFFFF", padx=20, pady=10, relief="raised",
                             bd=5, command=iniciar_analise)
    start_button.place(relx=0.5, rely=0.55, anchor='center')

    # Direitos autorais exibidos no final da tela
    copyright_label = tk.Label(root, text="Direitos Autorais © Arquindex.",
                               font=("Helvetica", 12, "bold"), fg="#FFFFFF", bg="#1C2833", padx=5, pady=5)
    copyright_label.place(relx=0.01, rely=0.95, anchor='w')

    # Iniciar o loop principal da interface gráfica
    root.mainloop()
#
class CircularProgressBar(ttk.Frame):
    def __init__(self, parent, size=100, thickness=10, max_value=100, *args, **kwargs):
        # Inicializa o Frame da classe pai (ttk.Frame)
        super().__init__(parent, *args, **kwargs)

        # Define os parâmetros da barra de progresso
        self.size = size  # Tamanho total do círculo
        self.thickness = thickness  # Espessura do arco da barra de progresso
        self.max_value = max_value  # Valor máximo para a barra de progresso
        self.value = 0  # Valor inicial da barra de progresso

        # Configura o Canvas para desenhar o círculo
        # Canvas é uma área retangular onde desenhamos elementos gráficos
        self.canvas = tk.Canvas(self, width=size, height=size, bg="black", highlightthickness=0)
        self.canvas.pack()

        # Adiciona o texto que mostra o percentual atual no centro da barra
        # O texto começa mostrando "0%" e está centralizado no meio do círculo
        self.text = self.canvas.create_text(size / 2, size / 2, text="0%", font=("Helvetica", 24), fill="white")

        # Desenha o arco da barra de progresso
        # O arco inicial é um círculo completo com extensão de 0 graus (ou seja, nada preenchido)
        self.arc = self.canvas.create_arc(
            self.thickness, self.thickness,
            self.size - self.thickness, self.size - self.thickness,
            start=90,  # Começa no topo (90 graus)
            extent=0,  # Extensão inicial do arco é 0
            outline="#00FF00",  # Cor do arco (verde)
            width=self.thickness,  # Espessura do arco
            style="arc"  # Estilo do arco (em vez de uma fatia cheia)
        )

    def set_value(self, value):
        """Define o valor atual da barra de progresso"""
        # Garante que o valor esteja entre 0 e o valor máximo permitido
        self.value = min(self.max_value, max(0, value))

        # Calcula a extensão do arco com base no valor atual
        # A extensão é uma fração de 360 graus que representa o valor atual
        extent = (self.value / self.max_value) * 360

        # Atualiza o arco no Canvas para representar o progresso atual
        # extent negativo para desenhar no sentido horário
        self.canvas.itemconfig(self.arc, extent=-extent)

        # Atualiza o texto no centro do círculo para mostrar a porcentagem atual
        # Converte o valor em uma porcentagem inteira e exibe no formato "XX%"
        self.canvas.itemconfig(self.text, text=f"{int((self.value / self.max_value) * 100)}%")
#
class PDFAnalyzerGUI:
    def __init__(self):
        # Inicializa os componentes da interface gráfica e outros atributos
        self.canvas = None
        self.open_folder_button = None
        self.open_analise_button = None
        self.pages_blank_after_ocr_label = None
        self.pages_total_checked_label = None
        self.circular_progress = None
        self.progress_label = None
        self.progress_var = None
        self.analyze_button = None
        self.select_label = None
        self.pages_low_info_label = None
        # Cria uma janela principal com um tema estilizado usando ThemedTk
        self.window = ThemedTk(theme="arc")  # Define o tema "arc" para a janela principal
        self.window.title("Analisador de PDFs - Digitalizados")  # Define o título da janela
        self.window.state("zoomed")  # Define a janela para ser iniciada em modo maximizado

        # Diretório selecionado e instâncias de classes auxiliares
        self.directory = None
        self.analyzer = PDFAnalyzer()  # Instância da classe PDFAnalyzer para realizar as análises
        self.report_generator = ReportGenerator()  # Instância da classe ReportGenerator para criar relatórios

        # Fila para gerenciar progresso de processamento
        self.progress_queue = queue.Queue()

        # Configurações de estilo e criação dos componentes da interface
        self.setup_style()  # Configuração dos estilos dos widgets
        self.create_widgets()  # Cria os widgets da interface

        # Inicia o processamento da fila de progresso
        self.process_queue()

        # Inicia a interface gráfica
        self.window.mainloop()

    def reset_for_new_analysis(self):
        """Reinicia a interface para uma nova análise."""
        # Limpa o diretório selecionado e redefine os componentes da interface
        self.directory = None
        self.reset_labels()  # Reseta as contagens dos rótulos
        self.select_label.config(text="Nenhum diretório selecionado")  # Atualiza o rótulo de seleção do diretório
        self.analyze_button.config(state="disabled")  # Desabilita o botão de análise
        self.canvas.delete("all")  # Limpa o canvas
        self.circular_progress.set_value(0)  # Reseta o progresso da barra circular
        self.progress_queue.queue.clear()  # Limpa a fila de progresso
        self.open_analise_button.config(state="disabled")  # Desabilita o botão de abrir análises

    def setup_style(self):
        # Configura o estilo dos componentes da interface gráfica
        style = ttk.Style(self.window)
        style.configure("TFrame", background="#0e0d2a")  # Azul mais escuro para melhor contraste
        style.configure("TLabel", background="#121212", foreground="white", font=("Segoe UI", 10))
        style.configure("Header.TLabel", foreground="#F0F0F0", font=("Segoe UI", 10, "bold"))
        style.configure("TButton", background="#374956", foreground="black", font=("Segoe UI", 10, "bold"), padding=10)
        style.map("TButton", background=[("active", "#516B7F")], relief=[("pressed", "ridge"), ("!pressed", "flat")])

    def create_widgets(self):
        # Criação dos dois frames para dividir a interface em colunas
        main_frame_left = ttk.Frame(self.window, padding="20")  # Frame esquerdo para controles e progresso
        main_frame_left.pack(side='left', fill='y', expand=False)

        main_frame_right = ttk.Frame(self.window, padding="20")  # Frame direito para exibir imagens
        main_frame_right.pack(side='right', fill='both', expand=True)

        # Componentes da coluna esquerda (controles e progresso)
        header_label = ttk.Label(main_frame_left, text="Selecione um diretório para análise de PDFs:",
                                 style="Header.TLabel")  # Rótulo de instrução
        header_label.pack(pady=(0, 20))

        # Botão para selecionar o diretório dos PDFs
        select_button = ttk.Button(main_frame_left, text="Selecionar Diretório", command=self.select_directory,
                                   width=30)
        select_button.pack(pady=10)

        # Label para exibir o diretório selecionado
        self.select_label = ttk.Label(main_frame_left, text="Nenhum diretório selecionado", foreground="#a6a6a6")
        self.select_label.pack(pady=(5, 15))

        # Botão para iniciar a análise dos PDFs
        self.analyze_button = ttk.Button(main_frame_left, text="Iniciar Análise", state="disabled",
                                         command=self.start_analysis, width=30)  # Desabilitado até que um diretório seja selecionado
        self.analyze_button.pack(pady=10)

        # Barra de progresso circular para exibir o progresso da análise
        self.circular_progress = CircularProgressBar(main_frame_left, size=200, thickness=20, max_value=100)
        self.circular_progress.pack(pady=20)

        # Labels para exibir informações sobre a análise das páginas
        self.pages_blank_after_ocr_label = ttk.Label(main_frame_left, text="Página em Branco / Pouca Informação: 0")
        self.pages_blank_after_ocr_label.pack(pady=5)
        self.pages_low_info_label = ttk.Label(main_frame_left, text="Necessidade de verificar: 0")
        self.pages_low_info_label.pack(pady=5)
        self.pages_total_checked_label = ttk.Label(main_frame_left, text="Total de Páginas Verificadas: 0")
        self.pages_total_checked_label.pack(pady=5)

        # Frame para organizar os botões na mesma linha
        button_frame = ttk.Frame(main_frame_left)
        button_frame.pack(pady=15)

        # Botão para abrir a pasta do relatório
        self.open_folder_button = ttk.Button(button_frame, text="Abrir Pasta do Relatório", command=self.open_folder,
                                             width=25)
        self.open_folder_button.grid(row=0, column=0, padx=5)

        # Botão para abrir as análises pendentes
        self.open_analise_button = ttk.Button(button_frame, text="Abrir Análises", command=self.open_analysis_screen,
                                              width=25, state="disabled")  # Desabilitado inicialmente
        self.open_analise_button.grid(row=0, column=1, padx=5)

        # Botão "Iniciar Nova Análise" para reiniciar a interface
        self.restart_button = ttk.Button(button_frame, text="Iniciar Nova Análise", command=self.reset_for_new_analysis,
                                         width=25)
        self.restart_button.grid(row=0, column=2, padx=5)

        # Frame do canvas para exibir imagens na coluna direita
        canvas_frame = ttk.Frame(main_frame_right, padding="10", borderwidth=2, relief="ridge", style="TFrame")
        canvas_frame.pack(expand=True, fill="both")

        # Canvas para exibir imagens das páginas PDF
        self.canvas = tk.Canvas(canvas_frame, bg="white")  # Canvas para renderizar as imagens das páginas
        self.canvas.pack(expand=True, fill="both")

    def select_directory(self):
        # Abre um diálogo para selecionar o diretório contendo os arquivos PDF
        self.directory = filedialog.askdirectory()
        if self.directory:
            # Atualiza a label com o diretório selecionado
            self.select_label.config(text=f"Diretório Selecionado: {self.directory}")
            self.analyze_button.config(state="normal")  # Habilita o botão de análise

    def start_analysis(self):
        # Inicia a análise dos PDFs no diretório selecionado em uma nova thread
        if self.directory:
            self.analyze_button.config(state="disabled")  # Desabilita o botão durante a análise
            threading.Thread(target=self.run_analysis_thread, daemon=True).start()  # Inicia uma nova thread para não bloquear a interface

    def run_analysis_thread(self):
        # Executa a análise dos PDFs e gera um relatório
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")  # Cria um timestamp para o nome do relatório
        output_xlsx = os.path.join(self.directory, f"analysis_report_{timestamp}.xlsx")
        self.analyze_pdfs_in_directory(output_xlsx)

    def analyze_pdfs_in_directory(self, output_xlsx):
        # Analisa todos os PDFs no diretório selecionado e gera um relatório
        self.reset_labels()  # Reinicia as contagens
        pdf_files = [os.path.join(self.directory, f) for f in os.listdir(self.directory) if f.lower().endswith('.pdf')]
        total_pages = sum(fitz.open(pdf_file).page_count for pdf_file in pdf_files)  # Calcula o total de páginas dos PDFs
        total_pages_processed = 0

        # Itera sobre cada arquivo PDF no diretório
        for pdf_file in pdf_files:
            pdf_name = os.path.basename(pdf_file)
            with fitz.open(pdf_file) as pdf_document:  # Abre o arquivo PDF
                for page_num in range(pdf_document.page_count):
                    # Carrega a página e converte para imagem
                    page = pdf_document.load_page(page_num)
                    pix = page.get_pixmap()
                    img = Image.open(io.BytesIO(pix.tobytes("png")))  # Converte a página para imagem usando PIL

                    # Realiza análise na página
                    status, white_pixel_percentage, ocr_performed, extracted_text = self.analyzer.analyze_page(img)
                    # Adiciona os resultados ao gerador de relatórios
                    self.report_generator.add_record(pdf_name, page_num + 1, status, white_pixel_percentage,
                                                     ocr_performed, extracted_text)

                    # Atualizar labels e progresso
                    total_pages_processed += 1
                    self.update_labels(total_pages_processed)  # Atualiza as labels com o progresso
                    progress_percentage = (total_pages_processed / total_pages) * 100  # Calcula a porcentagem do progresso
                    self.progress_queue.put(progress_percentage)  # Adiciona o progresso à fila

                    # Enviar a imagem para ser exibida no canvas
                    self.progress_queue.put(("image", img))

        # Finaliza o relatório após processar todas as páginas
        self.report_generator.finalize(output_xlsx)
        if not os.path.exists(output_xlsx):
            print("Erro: O relatório não foi criado.")
            return
        self.progress_queue.put("DONE")  # Indica que a análise foi concluída

    def process_queue(self):
        # Processa os itens da fila para atualizar a interface em tempo real
        try:
            while True:
                message = self.progress_queue.get_nowait()  # Tenta obter um item da fila
                if isinstance(message, tuple) and message[0] == "image":
                    image = message[1]
                    self.display_image_on_canvas(image)  # Exibe a imagem da página no canvas
                elif isinstance(message, float) or isinstance(message, int):
                    self.update_progress(message)  # Atualiza o progresso circular
                elif message == "DONE":
                    self.analyze_button.config(state="normal")  # Habilita o botão de análise
                    # Verifica se há páginas que necessitam de revisão
                    if self.analyzer.pages_blank_after_ocr_count == 0 and self.analyzer.pages_low_info_count == 0:
                        # Nenhuma página para revisar, desabilita o botão
                        self.open_analise_button.config(state="disabled")
                    else:
                        # Há páginas para revisar, habilita o botão
                        self.open_analise_button.config(state="normal")
                    messagebox.showinfo("Análise Concluída",
                                        "A análise foi concluída e o relatório foi gerado com sucesso!")
                    break  # Sai do loop ao concluir a análise
        except queue.Empty:
            pass
        # Verifica a fila novamente após 100 ms
        self.window.after(100, self.process_queue)

    def update_progress(self, value):
        """Atualiza o progresso circular"""
        self.circular_progress.set_value(value)  # Define o valor atual da barra de progresso circular

    def update_labels(self, total_pages_checked):
        # Atualiza as labels que exibem informações sobre a análise
        self.pages_blank_after_ocr_label.config(
            text=f"Página em Branco / Pouca Informação: {self.analyzer.pages_blank_after_ocr_count}")
        self.pages_total_checked_label.config(
            text=f"Total de Páginas Verificadas: {total_pages_checked}")
        self.pages_low_info_label.config(
            text=f"Necessário revisar: {self.analyzer.pages_low_info_count}")

    def display_image_on_canvas(self, image):
        print("Exibindo imagem no canvas...")

        # Obtenha dimensões do canvas e da imagem
        canvas_width = self.canvas.winfo_width()
        canvas_height = self.canvas.winfo_height()
        img_width, img_height = image.size

        # Define dimensões padrão se o canvas não estiver inicializado corretamente
        if canvas_width == 1 or canvas_height == 1:
            canvas_width, canvas_height = 400, 600
            self.canvas.config(width=canvas_width, height=canvas_height)

        # Define o filtro de reamostragem adequado
        try:
            resample_filter = Image.Resampling.LANCZOS  # Usa o filtro LANCZOS para melhor qualidade
        except AttributeError:
            resample_filter = Image.LANCZOS

        # Calcula a nova escala da imagem para caber no canvas
        ratio = min(canvas_width / img_width, canvas_height / img_height)
        new_size = (int(img_width * ratio), int(img_height * ratio))
        image = image.resize(new_size, resample=resample_filter)  # Redimensiona a imagem mantendo a proporção

        # Calcula a posição para centralizar a imagem no canvas
        x = (canvas_width - new_size[0]) // 2
        y = (canvas_height - new_size[1]) // 2
        print(f"Posicionando imagem no canvas: ({x}, {y})")

        # Exibe a imagem no canvas
        tk_image = ImageTk.PhotoImage(image)
        self.canvas.delete("all")  # Limpa o canvas antes de desenhar a nova imagem
        self.canvas.create_image(x, y, anchor="nw", image=tk_image)
        self.canvas.image = tk_image  # Evita que o Python faça coleta de lixo da imagem

    def open_folder(self):
        # Abre o diretório contendo os PDFs ou relatórios
        if self.directory:
            try:
                # Use platform.system() para verificar o sistema operacional corretamente
                current_system = platform.system()
                if current_system == "Windows":
                    os.startfile(self.directory)  # Abre a pasta no Windows
                elif current_system == "Darwin":
                    subprocess.Popen(["open", self.directory])  # Abre a pasta no macOS
                elif current_system == "Linux":
                    subprocess.Popen(["xdg-open", self.directory])  # Abre a pasta no Linux
                else:
                    raise EnvironmentError("Sistema operacional não suportado")
            except Exception as e:
                messagebox.showerror("Erro", f"Não foi possível abrir a pasta: {str(e)}")

    def open_analysis_screen(self):
        # Verifica se um diretório foi selecionado
        if not self.directory:
            messagebox.showerror("Erro", "Nenhum diretório selecionado.")
            return

        # Encontra o arquivo de relatório mais recente com o padrão analysis_report_YYYYMMDD_HHMMSS.xlsx
        report_files = [f for f in os.listdir(self.directory) if
                        f.startswith("analysis_report_") and f.endswith(".xlsx")]

        if not report_files:
            messagebox.showerror("Erro", "Nenhum relatório de análise encontrado.")
            return

        # Ordena os relatórios por data de criação em ordem decrescente para obter o mais recente
        report_files.sort(
            key=lambda x: datetime.strptime(x.replace("analysis_report_", "").replace(".xlsx", ""), "%Y%m%d_%H%M%S"),
            reverse=True)
        analysis_report_path = os.path.join(self.directory, report_files[0])

        # Minimiza a janela principal antes de abrir a nova tela
        self.window.iconify()

        # Cria a tela de análises pendentes
        AnalysisScreen(self.window, analysis_report_path)

    def reset_labels(self):
        """Zera as contagens de labels para uma nova análise."""
        self.pages_blank_after_ocr_label.config(text="Página em Branco / Pouca Informação: 0")
        self.pages_total_checked_label.config(text="Total de Páginas Verificadas: 0")
        self.pages_low_info_label.config(text="Necessidade de verificar: 0")

        # Resetando contadores da instância do PDFAnalyzer
        self.analyzer.pages_blank_after_ocr_count = 0
        self.analyzer.pages_low_info_count = 0
        self.analyzer.pages_blank_count = 0
#
class AnalysisScreen:
    def __init__(self, master, analysis_report_path):
        # Inicializa a tela de análise e configurações de estilo e layout
        print("Inicializando AnalysisScreen...")
        self.window = tk.Toplevel(master)  # Cria uma nova janela de nível superior
        self.window.title("Análises Pendentes")  # Define o título da janela
        self.window.state('zoomed')  # Maximiza a janela
        self.window.configure(bg="#2B3E50")  # Define a cor de fundo

        # Configurações de estilo para os componentes da interface
        style = ttk.Style(self.window)
        style.configure("TFrame", background="#0e0d2a")  # Cor de fundo do frame
        style.configure("TLabel", background="#2B3E50", foreground="white", font=("Segoe UI", 10))
        style.configure("Header.TLabel", foreground="#ECECEC", font=("Segoe UI", 12, "bold"))

        # Guarda o caminho do relatório e do diretório selecionado
        self.analysis_report_path = analysis_report_path
        self.selected_directory = os.path.dirname(analysis_report_path)

        # Configuração da interface (Frames, Labels, Treeview)
        self.header_frame = ttk.Frame(self.window, style="TFrame")  # Frame de cabeçalho
        self.header_frame.pack(pady=10, padx=20, fill='x')

        self.report_label = ttk.Label(
            self.header_frame,
            text=f"Relatório: {os.path.basename(self.analysis_report_path)}",
            style="Header.TLabel"
        )
        self.report_label.pack(side='left', padx=5)  # Exibe o nome do relatório

        self.date_label = ttk.Label(
            self.header_frame,
            text=f"Data: {pd.to_datetime('today').strftime('%d/%m/%Y')}",
            style="Header.TLabel"
        )
        self.date_label.pack(side='right', padx=5)  # Exibe a data atual

        # Treeview para exibir arquivos pendentes
        columns = ("Arquivo PDF", "Página", "Status")
        self.pending_files_tree = ttk.Treeview(
            self.window, columns=columns, show="headings", height=25
        )
        self.pending_files_tree.pack(pady=20, padx=10, fill='y', side='left')

        # Configura colunas e cabeçalhos
        self.pending_files_tree.heading("Arquivo PDF", text="Arquivo PDF")
        self.pending_files_tree.heading("Página", text="Página")
        self.pending_files_tree.heading("Status", text="Status")
        self.pending_files_tree.column("Arquivo PDF", width=200)
        self.pending_files_tree.column("Página", width=50, anchor='center')
        self.pending_files_tree.column("Status", width=190, anchor='center')

        # Frame para visualização de PDF
        self.pdf_view_frame = ttk.Frame(
            self.window, padding="10", borderwidth=2, relief="ridge", style="TFrame"
        )
        self.pdf_view_frame.pack(pady=10, padx=10, side='right', fill='both', expand=True)

        # Canvas para exibir a imagem do PDF
        self.pdf_canvas = tk.Canvas(self.pdf_view_frame, bg="#2B3E50")
        self.pdf_canvas.pack(expand=True, fill='both')
        self.pdf_canvas.bind('<Configure>', self.center_image)

        # Botão para deletar a página selecionada
        delete_button = tk.Button(
            self.pdf_view_frame,
            text="Deletar Página Selecionada",
            command=self.delete_selected_pdf,
            bg="#f44336",
            fg="white",
            font=("Segoe UI", 10, "bold"),
            activebackground="#e53935",
            padx=10,
            pady=5
        )
        delete_button.pack(side='bottom', pady=5)

        # Botão para abrir o diretório dos PDFs
        open_button = tk.Button(
            self.window,
            text="Abrir Diretório dos PDFs",
            command=self.open_pdf_directory
        )
        open_button.pack(pady=10)

        # Label de carregamento
        self.loading_label = tk.Label(
            self.window,
            text="",
            bg="#2B3E50",
            fg="white",
            font=("Segoe UI", 10, "bold")
        )
        self.loading_label.pack(pady=5)

        # Carrega os arquivos pendentes no Treeview
        self.load_pending_files()
        self.pending_files_tree.bind('<<TreeviewSelect>>', self.on_pdf_select)

    # Função para deletar PDF ao pressionar Delete
    def on_delete_key_press(self, event):
        self.delete_selected_pdf()

    # Função para carregar arquivos pendentes no Treeview
    def load_pending_files(self):
        self.pending_files_tree.delete(*self.pending_files_tree.get_children())  # Limpa o Treeview
        self.analyzed_pages = []  # Exemplo de uma lista para armazenar dados temporários de análise

        try:
            # Carrega o relatório de análise do Excel e insere no Treeview
            df = pd.read_excel(self.analysis_report_path)
            pending_pages = df[(df['Status'] != 'OK') & (df['Status'] != 'Identificado conteúdo após reanálise')][['Arquivo PDF', 'Página', 'Status']].drop_duplicates()
            for _, row in pending_pages.iterrows():
                pdf_name = row['Arquivo PDF']
                page_number = row['Página']
                status = row['Status']
                self.pending_files_tree.insert("", 'end', values=(pdf_name, page_number, status))
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar o relatório: {str(e)}")

    # Função para selecionar um PDF e renderizar a página no Canvas
    def on_pdf_select(self, event):
        selected_item = self.pending_files_tree.selection()
        if not selected_item:
            return
        try:
            selected_entry = self.pending_files_tree.item(selected_item, "values")
            pdf_name, page_info, status = selected_entry
            self.selected_pdf = os.path.join(self.selected_directory, pdf_name)
            page_number = int(page_info)  # Número da página (1-based)
            self.selected_page_index = page_number - 1  # Índice da página (0-based)

            if os.path.exists(self.selected_pdf):
                try:
                    self.render_pdf_page(self.selected_pdf, page_number)  # Renderiza a página do PDF
                except Exception as e:
                    messagebox.showerror("Erro", f"Não foi possível abrir o PDF: {str(e)}")
            else:
                messagebox.showerror("Erro", "Arquivo PDF não encontrado!")
        except (tk.TclError, ValueError) as e:
            print(f"Erro ao acessar o item: {str(e)}")

    # Função para abrir o diretório dos PDFs
    def open_pdf_directory(self):
        if os.path.exists(self.selected_directory):
            os.startfile(self.selected_directory)
        else:
            messagebox.showerror("Erro", "Diretório dos PDFs não encontrado!")

    # Função para confirmar a exclusão de uma página
    def confirm_delete_page(self, page_number):
        result = [False]  # Usamos uma lista para capturar o resultado
        confirm_win = tk.Toplevel()
        confirm_win.transient(self.window)
        confirm_win.grab_set()  # Torna a janela modal
        confirm_win.title("Confirmação")
        confirm_win.geometry("300x150")
        confirm_win.resizable(False, False)

        label = tk.Label(confirm_win, text=f"Deseja excluir a página {page_number}?")
        label.pack(pady=20)

        button_frame = tk.Frame(confirm_win)
        button_frame.pack(pady=10)

        def on_yes():
            result[0] = True
            confirm_win.destroy()

        def on_no():
            result[0] = False
            confirm_win.destroy()

        yes_button = tk.Button(button_frame, text="Sim", command=on_yes, bg="green", fg="white", width=10)
        yes_button.pack(side='left', padx=10)

        no_button = tk.Button(button_frame, text="Não", command=on_no, bg="red", fg="white", width=10)
        no_button.pack(side='right', padx=10)

        confirm_win.wait_window()
        return result[0]

    # Função para deletar a página selecionada
    def delete_selected_pdf(self):
        try:
            selected_items = self.pending_files_tree.selection()
            if not selected_items:
                messagebox.showwarning("Aviso", "Nenhum PDF selecionado!")
                return

            pages_to_delete = {}
            for item in selected_items:
                pdf_name, page_info, status = self.pending_files_tree.item(item, "values")
                page_index = int(page_info) - 1

                if not self.confirm_delete_page(int(page_info)):
                    continue

                if pdf_name not in pages_to_delete:
                    pages_to_delete[pdf_name] = []
                pages_to_delete[pdf_name].append(page_index)

            for pdf_name, page_indices in pages_to_delete.items():
                pdf_path = os.path.join(self.selected_directory, pdf_name)
                if os.path.exists(pdf_path):
                    pdf_document = fitz.open(pdf_path)

                    for page_index in sorted(page_indices, reverse=True):
                        if 0 <= page_index < pdf_document.page_count:
                            pdf_document.delete_page(page_index)

                    temp_pdf_path = pdf_path + ".tmp"
                    pdf_document.save(temp_pdf_path)
                    pdf_document.close()

                    os.remove(pdf_path)
                    os.rename(temp_pdf_path, pdf_path)

                    self.update_report_and_treeview(pdf_name, page_indices)
                    self.clear_canvas()
                else:
                    messagebox.showerror("Erro", f"O arquivo PDF {pdf_name} não foi encontrado.")

            messagebox.showinfo("Sucesso", "As páginas selecionadas foram deletadas com sucesso.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao deletar as páginas: {str(e)}")

    # Atualiza o relatório e a Treeview após exclusão de páginas
    def update_report_and_treeview(self, pdf_name, deleted_page_indices):
        report_path = self.analysis_report_path
        df = pd.read_excel(report_path)

        deleted_page_indices = sorted(deleted_page_indices)

        for page_index in deleted_page_indices:
            page_number = page_index + 1
            df = df[~((df['Arquivo PDF'] == pdf_name) & (df['Página'] == page_number))]

        for page_index in deleted_page_indices:
            page_number = page_index + 1
            df.loc[(df['Arquivo PDF'] == pdf_name) & (df['Página'] > page_number), 'Página'] -= 1

        df.to_excel(report_path, index=False)
        self.pending_files_tree.delete(*self.pending_files_tree.get_children())
        self.load_pending_files()

    # Renderiza uma página PDF no Canvas
    def render_pdf_page(self, pdf_path, page_number):
        try:
            os.environ["PDF2IMAGE_PDFIUM_PATH"] = r"C:/pdfium/pdfium.dll"
            from pdf2image import convert_from_path

            images = convert_from_path(pdf_path, first_page=page_number, last_page=page_number, dpi=200)
            if not images:
                messagebox.showerror("Erro", f"Número de página {page_number} está fora do intervalo.")
                return

            pil_image = images[0]
            canvas_width = self.pdf_canvas.winfo_width()
            canvas_height = self.pdf_canvas.winfo_height()
            image_ratio = pil_image.width / pil_image.height
            canvas_ratio = canvas_width / canvas_height

            if image_ratio > canvas_ratio:
                new_width = canvas_width
                new_height = int(canvas_width / image_ratio)
            else:
                new_height = canvas_height
                new_width = int(canvas_height * image_ratio)

            pil_image = pil_image.resize((new_width, new_height), Image.LANCZOS)
            self.pdf_image = ImageTk.PhotoImage(pil_image)

            self.pdf_canvas.delete("all")
            x = (canvas_width - new_width) // 2
            y = (canvas_height - new_height) // 2
            self.pdf_canvas.create_image(x, y, anchor="nw", image=self.pdf_image)
            self.pdf_canvas.image = self.pdf_image
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao renderizar a página do PDF: {str(e)}")

    # Limpa o Canvas
    def clear_canvas(self):
        self.pdf_canvas.delete("all")
        self.pdf_canvas.image = None

    # Centraliza a imagem no Canvas ao redimensionar
    def center_image(self, event):
        self.pdf_canvas.delete('all')
        if hasattr(self, 'pdf_image'):
            self.pdf_canvas.create_image(
                self.pdf_canvas.winfo_width() // 2,
                self.pdf_canvas.winfo_height() // 2,
                anchor='center',
                image=self.pdf_image,
                tags='pdf_image'
            )
#
class PDFAnalyzer:
    def __init__(self, min_text_length=10, pixel_threshold=0.989, language='eng+por'):
        # Inicializa o objeto PDFAnalyzer com parâmetros de configuração
        print("Inicializando PDFAnalyzer...")
        self.min_text_length = min_text_length  # Tamanho mínimo do texto para considerar OCR como bem-sucedido
        self.pixel_threshold = pixel_threshold  # Limite de pixels brancos para definir uma página como em branco
        self.language = language  # Idiomas usados no OCR

        # Contadores para rastrear status de análise das páginas
        self.pages_blank_count = 0  # Contador para páginas detectadas como em branco
        self.pages_blank_after_ocr_count = 0  # Contador para páginas em branco mesmo após OCR
        self.pages_ocr_analyzed_count = 0  # Contador para páginas reclassificadas após OCR
        self.total_characters = 0  # Total de caracteres encontrados
        self.correct_characters = 0  # Contagem de caracteres considerados corretos
        self.total_words = 0  # Total de palavras encontradas
        self.correct_words = 0  # Contagem de palavras corretas
        self.pages_low_info_count = 0  # Contador de páginas com pouca informação
        print("PDFAnalyzer inicializado com sucesso.")

    def is_blank_or_noisy(self, image):
        print("Verificando se a imagem é em branco ou ruidosa...")

        # Recorta 10% de cada lado horizontalmente para evitar bordas ruidosas
        width, height = image.size
        crop_percent = 0.10  # Percentual de corte nas bordas laterais
        left = int(width * crop_percent)  # Define o limite esquerdo do corte
        right = int(width * (1 - crop_percent))  # Define o limite direito do corte
        cropped_image = image.crop((left, 0, right, height))  # Realiza o corte
        print(f"Imagem cortada para remover bordas: {left}px à {right}px")

        # Converte a imagem cortada para escala de cinza
        gray_image = cv2.cvtColor(np.array(cropped_image), cv2.COLOR_RGB2GRAY)
        print("Imagem convertida para escala de cinza.")

        # Aplica binarização adaptativa para separar fundo de texto
        binary_image = cv2.adaptiveThreshold(
            gray_image, 255,
            cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
            cv2.THRESH_BINARY,
            blockSize=15,
            C=10
        )
        print("Binarização adaptativa aplicada.")

        # Calcula a porcentagem de pixels brancos na imagem binarizada
        white_pixel_percentage = np.mean(binary_image == 255)
        print(f"Proporção de pixels brancos: {white_pixel_percentage:.2%}")

        # Verifica se a porcentagem de pixels brancos ultrapassa o limite para definir como "em branco"
        is_blank = white_pixel_percentage >= self.pixel_threshold
        print(f"Imagem é em branco: {is_blank}")

        return is_blank, white_pixel_percentage, cropped_image

    def perform_ocr_and_reclassify(self, cropped_image):
        print("Iniciando o processo de OCR e reclassificação...")

        try:
            # Aplica filtro mediano para reduzir ruído na imagem
            cropped_image = cropped_image.filter(ImageFilter.MedianFilter(size=3))
            print("Filtro mediano aplicado para reduzir ruído.")

            # Aumenta o contraste e a nitidez da imagem para melhorar o OCR
            cropped_image = ImageEnhance.Contrast(cropped_image).enhance(3.0)
            print("Contraste da imagem aumentado.")
            cropped_image = ImageEnhance.Sharpness(cropped_image).enhance(2.5)
            print("Nitidez da imagem aumentada.")

            # Converte a imagem para preto e branco com DPI alto para o OCR
            with io.BytesIO() as output:
                cropped_image.save(output, format="PNG", dpi=(300, 300))
                output.seek(0)
                with Image.open(output) as image_dpi:
                    image_bw = image_dpi.convert('L')  # Converte para escala de cinza
                    image_bw = ImageEnhance.Contrast(image_bw).enhance(2.0)
                    # Limiar para converter para imagem binária (preto e branco)
                    image_bw = image_bw.point(lambda x: 0 if x < 140 else 255, '1')
                    print("Imagem convertida para preto e branco para OCR.")

                    # Configuração do Tesseract para OCR
                    custom_config = r'--oem 3 --psm 6'
                    text = pytesseract.image_to_string(image_bw, lang=self.language, config=custom_config)
                    print("OCR realizado com Tesseract.")

            # Limpa o texto extraído, removendo caracteres indesejados e múltiplos espaços
            text = re.sub(r'[^A-Za-z0-9À-ÿ\s]', ' ', text)
            text = re.sub(r'\s+', ' ', text).strip()
            print(f"Texto extraído pelo OCR: {text[:50]}... (truncado)" if len(
                text) > 50 else f"Texto extraído pelo OCR: {text}")

            # Define se o OCR foi bem-sucedido com base no comprimento do texto limpo
            ocr_successful = len(text) >= self.min_text_length
            print(f"OCR foi bem-sucedido: {ocr_successful}")
            return ocr_successful, text

        except pytesseract.TesseractError as e:
            print(f"Erro no OCR: {e}")
            return False, ""

    def analyze_page(self, img):
        # Analisa a imagem de uma página do PDF
        is_blank, white_pixel_percentage, cropped_img = self.is_blank_or_noisy(img)
        ocr_performed = False  # Indicador de OCR realizado
        status = "OK"  # Status inicial
        quantidade_caracteres = 0  # Contador de caracteres extraídos

        if is_blank:
            self.pages_blank_count += 1  # Incrementa contador de páginas em branco
            ocr_successful, extracted_text = self.perform_ocr_and_reclassify(cropped_img)
            ocr_performed = True  # OCR foi executado

            if extracted_text:
                quantidade_caracteres = len(extracted_text.replace(" ", ""))  # Conta caracteres não vazios

            # Define o status com base na análise de OCR e quantidade de caracteres extraídos
            if (ocr_successful or quantidade_caracteres >= 20) and white_pixel_percentage <= 0.995:
                status = "Identificado conteúdo após reanálise"
                self.pages_ocr_analyzed_count += 1
            elif quantidade_caracteres >= 2 and white_pixel_percentage >= self.pixel_threshold:
                status = "Necessidade de revisão"
                self.pages_low_info_count += 1  # Incrementa contador de páginas com pouca informação
            else:
                status = "Página em branco ou pouca info."
                self.pages_blank_after_ocr_count += 1  # Incrementa contador de páginas em branco após OCR

        return status, white_pixel_percentage, ocr_performed, quantidade_caracteres
#
class ReportGenerator:
    def __init__(self):
        # Inicializa o ReportGenerator criando uma nova planilha de trabalho
        print("Inicializando ReportGenerator...")
        self.wb = Workbook()  # Cria um novo arquivo do Excel (Workbook)
        self.ws = self.wb.active  # Define a planilha ativa
        self.ws.title = "PDF Analysis Report"  # Nomeia a planilha ativa
        print("Workbook e Worksheet inicializados.")

        # Define os cabeçalhos das colunas, incluindo os novos campos 'OCR Feito' e 'Texto Extraído'
        self.headers = ["Arquivo PDF", "Página", "Status", "Porcentagem de Pixels Brancos", "OCR Feito",
                        "Texto Extraído"]
        self.ws.append(self.headers)  # Adiciona os cabeçalhos como primeira linha da planilha
        print(f"Worksheet inicializada com cabeçalhos: {self.headers}")

    def add_record(self, pdf_name, page_num, status, white_pixel_percentage, ocr_performed, extracted_text):
        try:
            # Insere um registro (linha) com as informações da análise de uma página
            print(
                f"Adicionando registro: PDF Name={pdf_name}, Página={page_num}, Status={status}, Porcentagem de Pixels Brancos={white_pixel_percentage}")

            # Converte o valor de `ocr_performed` para texto 'Sim' ou 'Não'
            ocr_feito = 'Sim' if ocr_performed else 'Não'

            # Cria uma linha com os dados formatados
            row = [
                pdf_name,
                page_num,
                status,
                f"{white_pixel_percentage:.2%}",  # Converte para porcentagem com duas casas decimais
                ocr_feito,
                extracted_text
            ]
            self.ws.append(row)  # Adiciona a linha à planilha
            print(f"Linha adicionada na planilha: {row}")

            # Destaca a linha em vermelho se o status for "Precisa de Atenção"
            if status == "Precisa de Atenção":
                print("Status 'Precisa de Atenção' detectado. Destacando a linha em vermelho.")
                fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                for col_idx in range(1, len(row) + 1):
                    cell = self.ws.cell(row=self.ws.max_row, column=col_idx)  # Acessa cada célula da linha
                    cell.fill = fill  # Aplica o preenchimento em vermelho
                    print(f"Célula {cell.coordinate} destacada em vermelho.")

            print("Registro adicionado com sucesso.")
        except Exception as e:
            print(f"Erro ao adicionar registro: {e}")

    def finalize(self, output_path):
        print("Finalizando o relatório...")

        # Verifica se o diretório de destino existe, caso contrário, cria o diretório
        dir_path = os.path.dirname(output_path)
        print(f"Verificando existência do diretório: {dir_path}")
        if not os.path.exists(dir_path):
            try:
                os.makedirs(dir_path)  # Cria o diretório
                print(f"Diretório criado: {dir_path}")
            except OSError as e:
                print(f"Erro ao criar o diretório: {e}")
                return
        else:
            print(f"Diretório já existe: {dir_path}")

        # Aplica estilo de tabela e ajusta a largura das colunas
        try:
            max_row = self.ws.max_row
            max_col = self.ws.max_column
            table_ref = f"A1:{get_column_letter(max_col)}{max_row}"  # Define a referência para a tabela
            print(f"Criando tabela com referência: {table_ref}")

            # Gera um nome único para a tabela com base no timestamp atual
            table_name = f"PDFAnalysisTable_{int(datetime.now().timestamp())}"

            # Cria e estiliza a tabela
            tab = Table(displayName=table_name, ref=table_ref)
            style = TableStyleInfo(
                name="TableStyleMedium9",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=True
            )
            tab.tableStyleInfo = style
            self.ws.add_table(tab)  # Adiciona a tabela à planilha
            print("Tabela criada com estilo aplicado.")

            # Ajusta a largura de cada coluna para melhorar a legibilidade
            for col in self.ws.columns:
                max_length = 0  # Define o comprimento máximo inicial
                column = col[0].column  # Obtém o índice da coluna
                column_letter = get_column_letter(column)  # Converte o índice para letra
                print(f"Ajustando largura da coluna {column_letter}...")

                # Percorre as células para calcular o comprimento máximo
                for cell in col:
                    cell_value_length = len(str(cell.value))  # Comprimento do valor da célula
                    print(f"Célula {cell.coordinate} valor: '{cell.value}', comprimento: {cell_value_length}")
                    max_length = max(max_length, cell_value_length)

                # Ajusta a largura com base no comprimento máximo dos dados da coluna
                adjusted_width = max_length + 2
                self.ws.column_dimensions[column_letter].width = adjusted_width
                print(f"Largura da coluna {column_letter} ajustada para {adjusted_width}.")

            # Salva o relatório no caminho especificado
            print(f"Salvando o relatório no caminho: {output_path}")
            self.wb.save(output_path)
            print(f"Relatório salvo em: {output_path}")

        except Exception as e:
            print(f"Erro ao salvar o relatório: {e}")
            messagebox.showerror("Erro", f"Erro ao salvar o relatório: {str(e)}")  # Exibe uma mensagem de erro
#
class TesseractConfig:
    def __init__(self, tessdata_path, tesseract_cmd):
        # Inicializa o objeto TesseractConfig e configura os caminhos do Tesseract OCR
        self.tessdata_prefix = tessdata_path  # Caminho para os arquivos de dados do Tesseract (tessdata)
        self.tesseract_cmd = tesseract_cmd  # Caminho para o executável do Tesseract OCR
        self.configure_tesseract()  # Configura as variáveis de ambiente e o comando Tesseract

    def configure_tesseract(self):
        # Configura as variáveis de ambiente para o Tesseract OCR
        os.environ['TESSDATA_PREFIX'] = self.tessdata_prefix  # Define o caminho do tessdata no ambiente
        pytesseract.tesseract_cmd = self.tesseract_cmd  # Define o caminho do executável Tesseract para o pytesseract

    def test_setup(self):
        # Testa a configuração para garantir que o Tesseract OCR foi configurado corretamente
        try:
            print("Verificando a configuração do Tesseract OCR...")

            # Verifica se o TESSDATA_PREFIX foi configurado e se é um diretório válido
            tessdata_prefix_env = os.environ.get('TESSDATA_PREFIX')
            if not tessdata_prefix_env or not os.path.isdir(tessdata_prefix_env):
                raise EnvironmentError("TESSDATA_PREFIX não está configurado corretamente.")

            # Obtém e exibe a versão do Tesseract para confirmar que ele está acessível
            tesseract_version = pytesseract.get_tesseract_version()
            print(f"Tesseract OCR instalado corretamente. Versão: {tesseract_version}")

            # Mostra uma mensagem de sucesso usando messagebox
            messagebox.showinfo("Tesseract OCR", f"Tesseract OCR instalado corretamente.\nVersão: {tesseract_version}")

        except Exception as e:
            # Captura erros na configuração e exibe uma mensagem de erro
            print(f"Erro ao inicializar o Tesseract OCR: {e}")
            messagebox.showerror("Erro Tesseract OCR", f"Erro ao inicializar o Tesseract OCR.\n{str(e)}")
            sys.exit(1)  # Encerra o programa caso haja um erro crítico


if __name__ == "__main__":
    print("Executando main.py como script principal")
    iniciar_interface_principal()
